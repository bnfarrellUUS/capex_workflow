"""Excel export of the requests list (openpyxl).

Visibility comes from request_service.list_requests; `q` replicates the
client-side list search so the file matches the filtered screen.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services import request_service

_HEADER_FILL = PatternFill("solid", start_color="0B2A4A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MONEY_FMT = '"$"#,##0.00'
_DATE_FMT = "yyyy-mm-dd"


def _division_name(req):
    return f"{req.division.number} — {req.division.name}" if req.division else None


def _yn(value):
    return "Yes" if value else "No"


def _day(value):
    # Excel can't store tz-aware datetimes; the date is all we report anyway.
    return value.date() if value else None


def matches_query(req, q):
    """Case-insensitive contains over number, division display name, and
    requestor name — same fields as the frontend list search."""
    if not q or not q.strip():
        return True
    needle = q.strip().lower()
    fields = (req.number, _division_name(req),
              req.requestor.name if req.requestor else None)
    return any(needle in f.lower() for f in fields if f)


# (header, getter, kind); kind picks the cell number format.
COLUMNS = [
    ("Number", lambda r: r.number, "text"),
    ("Status", lambda r: r.status, "text"),
    ("Division", _division_name, "text"),
    ("Requestor", lambda r: r.requestor.name if r.requestor else None, "text"),
    ("Request Date", lambda r: _day(r.request_date), "date"),
    ("Total Cost", lambda r: r.total_cost, "money"),
    ("Current Level", lambda r: r.current_level, "num"),
    ("Required Levels", lambda r: r.required_levels, "num"),
    ("Budgeted", lambda r: _yn(r.budgeted), "text"),
    ("Replacement", lambda r: _yn(r.replacement), "text"),
    ("Health & Safety", lambda r: _yn(r.health_safety), "text"),
    ("Revenue Generating", lambda r: _yn(r.revenue_generating), "text"),
    ("Environmental", lambda r: _yn(r.environmental), "text"),
    ("Competitive Bids", lambda r: _yn(r.competitive_bids), "text"),
    ("Lease Recommended", lambda r: _yn(r.lease_recommended), "text"),
    ("Asset Life", lambda r: r.asset_life, "text"),
    ("IRR After Tax", lambda r: r.irr_after_tax, "num"),
    ("First-Year EBIT", lambda r: r.first_year_ebit, "money"),
    ("Annual Savings", lambda r: r.annual_savings, "money"),
    ("Payback Years", lambda r: r.payback_years, "num"),
    ("NPV Savings", lambda r: r.npv_savings, "money"),
    ("Asset Number", lambda r: r.asset_number, "text"),
    ("GL Account", lambda r: r.gl_account, "text"),
    ("Useful Life (Years)", lambda r: r.useful_life_years, "num"),
    ("Useful Life (Months)", lambda r: r.useful_life_months, "num"),
    ("In-Service Date", lambda r: _day(r.in_service_date), "date"),
    ("Cost: Autos & Trucks", lambda r: r.cost_autos_trucks, "money"),
    ("Cost: Machinery & Equipment", lambda r: r.cost_machinery, "money"),
    ("Cost: Improvements", lambda r: r.cost_improvements, "money"),
    ("Cost: Furniture & Fixtures", lambda r: r.cost_furniture, "money"),
    ("Cost: IT / Computer", lambda r: r.cost_it_computer, "money"),
    ("Cost: Miscellaneous", lambda r: r.cost_misc, "money"),
    ("Finance Completed", lambda r: _yn(r.finance_completed), "text"),
]


def build_workbook(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CAPEX Requests"
    ws.append([header for header, _, _ in COLUMNS])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"
    for req in rows:
        ws.append([getter(req) for _, getter, _ in COLUMNS])
        for idx, (_, _, kind) in enumerate(COLUMNS, start=1):
            if kind == "money":
                ws.cell(row=ws.max_row, column=idx).number_format = _MONEY_FMT
            elif kind == "date":
                ws.cell(row=ws.max_row, column=idx).number_format = _DATE_FMT
    for idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 18
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_xlsx(viewer, scope="mine", status=None, division_id=None, q=None) -> bytes:
    rows = request_service.list_requests(
        viewer, scope=scope, status=status, division_id=division_id)
    rows = [r for r in rows if matches_query(r, q)]
    rows.sort(key=lambda r: r.number or "")
    return build_workbook(rows)
