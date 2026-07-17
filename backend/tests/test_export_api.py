from io import BytesIO

from openpyxl import load_workbook

from tests.factories import make_user, make_division, make_draft


def _login(client, user):
    client.post("/api/auth/login",
                json={"email": user.email, "password": "secret123"})


def _sheet(res):
    assert res.status_code == 200
    assert res.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return load_workbook(BytesIO(res.data)).active


def _numbers(ws):
    return [row[0].value for row in ws.iter_rows(min_row=2)]


def test_export_requires_login(client, app):
    assert client.get("/api/requests/export.xlsx").status_code == 401


def test_export_mine_only_own_rows(client, app):
    a = make_user("a", roles='["REQUESTOR"]')
    b = make_user("b", roles='["REQUESTOR"]')
    div = make_division()
    make_draft(a.id, div.id, number="CX000001")
    make_draft(b.id, div.id, number="CX000002")
    _login(client, a)
    ws = _sheet(client.get("/api/requests/export.xlsx?scope=mine"))
    assert _numbers(ws) == ["CX000001"]


def test_export_scope_all_denied_for_plain_requestor(client, app):
    a = make_user("a", roles='["REQUESTOR"]')
    b = make_user("b", roles='["REQUESTOR"]')
    div = make_division()
    make_draft(a.id, div.id, number="CX000001")
    make_draft(b.id, div.id, number="CX000002")
    _login(client, a)
    # Same fallback as the list route: non-ADMIN/FINANCE gets own rows.
    ws = _sheet(client.get("/api/requests/export.xlsx?scope=all"))
    assert _numbers(ws) == ["CX000001"]


def test_export_scope_all_for_finance_with_filters(client, app):
    fin = make_user("fin", roles='["FINANCE"]')
    a = make_user("a", roles='["REQUESTOR"]')
    b = make_user("b", roles='["REQUESTOR"]')
    div = make_division()
    make_draft(a.id, div.id, number="CX000001")
    make_draft(b.id, div.id, number="CX000002")
    _login(client, fin)
    ws = _sheet(client.get("/api/requests/export.xlsx?scope=all"))
    assert _numbers(ws) == ["CX000001", "CX000002"]
    ws = _sheet(client.get("/api/requests/export.xlsx?scope=all&q=CX000002"))
    assert _numbers(ws) == ["CX000002"]
    ws = _sheet(client.get(
        "/api/requests/export.xlsx?scope=all&status=APPROVED"))
    assert _numbers(ws) == []


def test_export_sets_attachment_filename(client, app):
    a = make_user("a", roles='["REQUESTOR"]')
    _login(client, a)
    res = client.get("/api/requests/export.xlsx")
    disp = res.headers["Content-Disposition"]
    assert disp.startswith("attachment; filename=capex-requests-")
    assert disp.endswith(".xlsx")
