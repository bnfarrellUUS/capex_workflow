from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.extensions import db
from app.services import export_service
from tests.factories import make_user, make_division, make_draft


def _ns_req(number="CX000001", division=None, requestor="Alice"):
    return SimpleNamespace(
        number=number,
        division=division,
        requestor=SimpleNamespace(name=requestor) if requestor else None,
    )


def test_matches_query_blank_matches_everything():
    assert export_service.matches_query(_ns_req(), None)
    assert export_service.matches_query(_ns_req(), "   ")


def test_matches_query_number_division_requestor_case_insensitive():
    div = SimpleNamespace(number="200", name="Field Services")
    req = _ns_req(number="CX000042", division=div, requestor="Bob Smith")
    assert export_service.matches_query(req, "cx0000")
    assert export_service.matches_query(req, "field serv")
    assert export_service.matches_query(req, "200 — field")
    assert export_service.matches_query(req, "bob")
    assert not export_service.matches_query(req, "zzz")


def test_matches_query_handles_missing_division_and_requestor():
    req = _ns_req(division=None, requestor=None)
    assert export_service.matches_query(req, "cx000001")
    assert not export_service.matches_query(req, "field")


def test_build_workbook_headers_and_types(app):
    user = make_user("req", roles='["REQUESTOR"]')
    div = make_division()
    r = make_draft(user.id, div.id, number="CX000009")
    r.total_cost = Decimal("30000")
    r.request_date = datetime(2026, 3, 5)
    r.budgeted = True
    r.cost_machinery = Decimal("12345.67")
    db.session.commit()

    data = export_service.export_xlsx(user, scope="mine")
    ws = load_workbook(BytesIO(data)).active

    headers = [c.value for c in ws[1]]
    assert headers[:6] == ["Number", "Status", "Division", "Requestor",
                           "Request Date", "Total Cost"]
    assert "Cost: Machinery & Equipment" in headers
    assert "Finance Completed" in headers

    row = list(ws[2])
    by_header = dict(zip(headers, row))
    assert by_header["Number"].value == "CX000009"
    assert by_header["Division"].value == "100 — Field Services"
    assert float(by_header["Total Cost"].value) == 30000.0
    assert by_header["Total Cost"].number_format == '"$"#,##0.00'
    assert by_header["Request Date"].value.year == 2026
    assert by_header["Budgeted"].value == "Yes"
    assert by_header["Replacement"].value == "No"
    assert float(by_header["Cost: Machinery & Equipment"].value) == 12345.67


def test_export_xlsx_filters_by_query_and_sorts_by_number(app):
    user = make_user("req", roles='["REQUESTOR"]')
    div = make_division()
    make_draft(user.id, div.id, number="CX000002")
    make_draft(user.id, div.id, number="CX000001")
    data = export_service.export_xlsx(user, scope="mine")
    ws = load_workbook(BytesIO(data)).active
    numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert numbers == ["CX000001", "CX000002"]

    data = export_service.export_xlsx(user, scope="mine", q="CX000002")
    ws = load_workbook(BytesIO(data)).active
    numbers = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert numbers == ["CX000002"]


def test_export_xlsx_empty_is_valid_workbook_with_header(app):
    user = make_user("req", roles='["REQUESTOR"]')
    data = export_service.export_xlsx(user, scope="mine")
    ws = load_workbook(BytesIO(data)).active
    assert ws[1][0].value == "Number"
    assert list(ws.iter_rows(min_row=2)) == []


def test_build_workbook_neutralizes_formula_injection(app):
    user = make_user("req", roles='["REQUESTOR"]')
    div = make_division()
    r = make_draft(user.id, div.id, number="CX000010")
    r.asset_life = "=1+1"
    r.gl_account = '=HYPERLINK("x")'
    db.session.commit()

    data = export_service.export_xlsx(user, scope="mine")
    ws = load_workbook(BytesIO(data)).active

    headers = [c.value for c in ws[1]]
    by_header = dict(zip(headers, list(ws[2])))

    asset_life_cell = by_header["Useful / Asset Life"]
    assert asset_life_cell.data_type == "s"
    assert asset_life_cell.value == "=1+1"

    gl_account_cell = by_header["GL Account"]
    assert gl_account_cell.data_type == "s"
    assert gl_account_cell.value == '=HYPERLINK("x")'
